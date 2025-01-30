from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth import authenticate,login,update_session_auth_hash
from django.views.generic import TemplateView,CreateView,UpdateView,DeleteView,DetailView
from django.contrib.auth.models import Group
from django.contrib.auth.models import User
from .forms import RegisterForm
from django.views import View
from .forms import RegisterForm,UserForm,ProfileForm,CourseForm,UserCreationForm,EditMateriaForm
from .models import Course,Registration,Materia,Mark,GlobalConfig 
from django.urls import reverse_lazy,reverse
from django.contrib import messages
from django.contrib.auth.mixins import UserPassesTestMixin,LoginRequiredMixin
import os
from django.conf import settings
from django.http import JsonResponse,HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.db.models import Q
from django.contrib.auth.views import PasswordChangeView,LoginView
from accounts.models import Profile
from django.template.loader import get_template
from weasyprint import HTML
from django.contrib.staticfiles.storage import staticfiles_storage
import random
import string
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import render_to_string
import openpyxl
from openpyxl.styles import PatternFill, Font
from django.utils import timezone
from .models import CITeacher, TeacherAttendance,Course,Materia
from .forms import AttendanceForm
from django.views.generic import ListView
import csv
from django.db.models import Q
from django.contrib.auth.hashers import make_password
from datetime import datetime
from django.templatetags.static import static
from .models import Schedule
from django.http import HttpResponseRedirect
from django.http import JsonResponse
from django.core.files.base import ContentFile
from weasyprint import HTML
from io import BytesIO
import base64
from .models import StudentHistory, User
from django.db import transaction
from datetime import date
from django.db.models import Avg
from django.db import transaction
from .models import Book
from .forms import BookForm



# FUNCION PARA CONVERTIR EL PLURAL DE UN GRUPO A SU SINGULAR
def plural_to_singular(plural):
    # Diccionario de palabras
    plural_singular = {
        "estudiantes": "estudiante",
        "profesores": "profesor",
        "secretarias": "secretaria",
        "administrativos": "administrativo",
    }
    return plural_singular.get(plural, "error")
# OBTENER COLOR Y GRUPO DE UN USUARIO
def get_group_and_color(user):
    group = user.groups.first()
    group_id = None
    group_name = None
    group_name_singular = None
    color = None

    if group:
        if group.name == 'estudiantes':
            color = 'bg-primary'
        elif group.name == 'profesores':
            color = 'bg-success'
        elif group.name == 'secretarias':
            color = 'bg-secondary'
        elif group.name == 'administrativos':
            color = 'bg-danger'

        group_id = group.id
        group_name = group.name
        group_name_singular = plural_to_singular(group.name)

    return group_id,group_name, group_name_singular, color

# DECORADOR PERSONALIZADO
def add_group_name_to_context(view_class):
    original_dispatch = view_class.dispatch

    def dispatch(self, request, *args, **kwargs):
        user = self.request.user
        group_id,group_name, group_name_singular, color = get_group_and_color(user)

        context = {
            'group_name': group_name,
            'group_name_singular': group_name_singular,
            'color': color
        }
        self.extra_context = context
        return original_dispatch(self, request, *args, **kwargs)
    view_class.dispatch = dispatch
    return view_class
# Create your views here.

# pagna de inicio
@add_group_name_to_context
class HomeView(TemplateView):
    template_name = 'home.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_time'] = datetime.now()  # Añadir la fecha y hora actual al contexto
        return context
#pagina de precios
@add_group_name_to_context
class PricingView(TemplateView):
    template_name ='pricing.html'

#registro de usuarios
@add_group_name_to_context
class RegisterView(View):
    def get(self, request):
        data = {
            'form': RegisterForm()
        }
        return render(request, 'registration/register.html', data)
    def post(self, request):
        user_creation_form = RegisterForm(data=request.POST)
        if user_creation_form.is_valid():
            user_creation_form.save()
            user = authenticate(username=user_creation_form.cleaned_data['username'],
                                password=user_creation_form.cleaned_data['password1'])
            login(request, user)
            # Actualizar el campo created_by_admin del modelo Profile
            #user.profile.created_by_admin = False
           #user.profile.save()

            return redirect('home')
        data = {
            'form': user_creation_form
        }
        return render(request, 'registration/register.html', data)
#pagina de perfil


@add_group_name_to_context
class ProfileView(TemplateView):
    template_name = 'profile/profile.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Obtener la configuración global para permitir agregar notas
        config = GlobalConfig.objects.first()
        context['allow_add_notes'] = config.allow_add_notes if config else True
        context['user_form'] = UserForm(instance=user)
        context['profile_form'] = ProfileForm(instance=user.profile)

        # Capturar el parámetro de búsqueda (si existe)
        search_query = self.request.GET.get('search', '')

        # Filtrar dependiendo del grupo del usuario
        if user.groups.filter(name='profesores').exists():
            context['group_name'] = 'profesores'
            assigned_materias = Materia.objects.filter(teacher=user).order_by('name')
            context['assigned_materias'] = assigned_materias

        elif user.groups.filter(name='estudiantes').exists():
            context['group_name'] = 'estudiantes'
            student_id = user.id
            registrations = Registration.objects.filter(student=user)
            enrolled_materias_set = set()
            for registration in registrations:
                materias = registration.course.materias.all()
                enrolled_materias_set.update(materias)
            enrolled_materias = list(enrolled_materias_set)
            context['student_id'] = student_id
            context['enrolled_materias'] = enrolled_materias

        elif user.groups.filter(name='secretarias').exists():
            all_courses = Course.objects.prefetch_related('materias').all()
            context['all_courses'] = all_courses

        elif user.groups.filter(name='administrativos').exists():
            context['group_name'] = 'administrativos'
            # Filtrar usuarios según la búsqueda
            all_users = User.objects.all()
            if search_query:
                all_users = all_users.filter(
                    Q(first_name__icontains=search_query) |
                    Q(last_name__icontains=search_query) |
                    Q(username__icontains=search_query)
                )

            all_groups = Group.objects.all()
            user_profiles = []
            for user in all_users:
                profile = user.profile
                user_groups = user.groups.all()
                user_profiles.append({
                    'user': user,
                    'groups': user_groups,
                    'profile': profile
                })

            all_courses = Course.objects.all()
            context['all_courses'] = all_courses
            context['user_profiles'] = user_profiles
            context['all_groups'] = all_groups

        return context

    def post(self, request, *args, **kwargs):
        user = self.request.user
        user_form = UserForm(request.POST, instance=user)
        profile_form = ProfileForm(request.POST, request.FILES, instance=user.profile)

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            return redirect('profile')

        context = self.get_context_data()
        context['user_form'] = user_form
        context['profile_form'] = profile_form
        return self.render_to_response(context)

        
# MOSTRAR TODOS LOS CURSOS
@add_group_name_to_context
class CoursesView(UserPassesTestMixin,TemplateView):
    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name='secretarias').exists()
    def handle_no_permission(self):
        return redirect('error')
    template_name = 'courses.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        courses = Course.objects.all().order_by('-id')
        context['courses'] = courses
        return context


# PAGINA DE ERROR DE ACCESO A PAGINA NO PERMITIDA
@add_group_name_to_context
class ErrorView(TemplateView):
    template_name = 'error.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        error_image_path = os.path.join(settings.MEDIA_URL, 'error.png')
        context['error_image_path'] = error_image_path
        return context
#crear curso por interfas
@add_group_name_to_context
class CourseCreateView( UserPassesTestMixin,CreateView):
    model = Course
    form_class = CourseForm
    template_name = 'create_course.html'
    #con esto volvemos a la vista courses
    success_url = reverse_lazy('courses')
    
    def test_func(self):
        #return self.request.user.groups.filter(name='administrativos', 'secretarias').exists()
        return self.request.user.groups.filter(name__in=['administrativos', 'secretarias']).exists()

    def handle_no_permission(self):
         return redirect('error')
    def form_valid(self, form):
        messages.success(self.request, 'El registro se ha guardado correctamente')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Ha ocurrido un error al guardar el registro')
        return self.render_to_response(self.get_context_data(form=form))

# EDICION DE UN CURSO
@add_group_name_to_context
class CourseEditView(UserPassesTestMixin, UpdateView):
    model = Course
    form_class = CourseForm
    template_name = 'edit_course.html'
    success_url = reverse_lazy('courses')

    def test_func(self):
        return self.request.user.groups.filter(name__in=['administrativos', 'secretarias']).exists()

    def handle_no_permission(self):
        return redirect('error')

    def form_valid(self, form):
        form.save()
        messages.success(self.request, 'El registro se ha actualizado satisfactoriamente')
        return redirect(self.success_url)

    def form_invalid(self, form):
        messages.error(self.request, 'Ha ocurrido un error al actualizar el registro')
        return self.render_to_response(self.get_context_data(form=form))
    
 # ELIMINACION DE UN curso
@add_group_name_to_context
class CourseDeleteView(UserPassesTestMixin, DeleteView):
    model = Course
    template_name = 'delete_course.html'
    success_url = reverse_lazy('courses')

    def test_func(self):
        return self.request.user.groups.filter(name='administrativos').exists()

    def handle_no_permission(self):
        return redirect('error')

    def form_valid(self, form):
        messages.success(self.request, 'El registro se ha eliminado correctamente')
        return super().form_valid(form)


# REGISTRO DE UN USUARIO EN UN CURSO
@add_group_name_to_context
class CourseEnrollmentView(TemplateView):
    def get(self, request, course_id):
        course = get_object_or_404(Course, id=course_id)

        if request.user.is_authenticated and request.user.groups.first().name == 'estudiantes':
            student = request.user

            # Crear un registro de inscripción asociado al estudiante y el curso
            registration = Registration(course=course, student=student)
            registration.save()

            messages.success(request, 'Inscripción exitosa')
        else:
            messages.error(request, 'No se pudo completar la inscripción')

        return redirect('courses')
    
#mostarar lista de alumnos y notas a los profesores
@add_group_name_to_context
class StudentListMarkView(LoginRequiredMixin,UserPassesTestMixin, TemplateView):
    template_name = 'student_list_mark.html'
    def test_func(self):
        user = self.request.user
        return user.groups.filter(name='profesores').exists()
    def handle_no_permission(self):
        return redirect('error')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        materia_id = self.kwargs['materia_id']
        materia = get_object_or_404(Materia, id=materia_id)
       
        context['grading_enabled'] = materia.activate_grading
        
      # Obtener las notas y los estudiantes relacionados, ordenados por apellido
        marks = Mark.objects.filter(materia=materia).order_by('student__last_name')

        student_data = {}
        for mark in marks:
            student = get_object_or_404(User, id=mark.student_id)
            if mark.student_id not in student_data:
                student_data[mark.student_id] = {
                    'mark_id': mark.id,
                    #'name': student.get_full_name(),
                    'last_name': student.last_name,
                    'first_name': student.first_name,
                    'mark_1': mark.mark_1,
                    'mark_2': mark.mark_2,
                    'mark_3': mark.mark_3,
                    'average': mark.average,
                }

        context['materia'] = materia
        context['student_data'] = list(student_data.values()) 
        return context
    def dispatch(self, request, *args, **kwargs):
        config = GlobalConfig.get_solo_config()
        print(f"allow_add_notes: {config.allow_add_notes}") 
        if not config.allow_add_notes:
            #print("Redirigiendo a la vista de error") Verificar si entra en esta lógica
            messages.error(request, 'No está permitido agregar notas en este momento.')
            return redirect(reverse('SinPermiso')) 
        return super().dispatch(request, *args, **kwargs)

#sin permiso profesores
@add_group_name_to_context
class SinPermisoView(TemplateView):
    template_name = 'error2.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        error_image_path = os.path.join(settings.MEDIA_URL, 'error.png')
        context['error_image_path'] = error_image_path
        return context
    #actualizar notas de alumnos
@add_group_name_to_context
class UpdateMarkView(UpdateView):
    model = Mark
    fields = ['mark_1', 'mark_2', 'mark_3']
    template_name = 'update_mark.html'
    
    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name='profesores').exists()
    def handle_no_permission(self):
        return redirect('error')
    def get_success_url(self):
        return reverse_lazy('student_list_mark', kwargs={'materia_id': self.object.materia.id})

    def form_valid(self, form):
        response = super().form_valid(form)
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mark = self.get_object()
        context['materia_name'] = mark.materia.name
        return context

    def get_object(self, queryset=None):
        mark_id = self.kwargs['mark_id']
        return get_object_or_404(Mark, id=mark_id)
#atedendace view per sin la parte de la asistencia
@add_group_name_to_context
class ListEstudent(TemplateView):
    template_name = 'student_2_list.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        materia_id = self.kwargs['materia_id']
        materia = get_object_or_404(Materia, id=materia_id)

        # Filtra las marcas de la materia y utiliza un set para evitar duplicados
        marks = Mark.objects.filter(materia=materia).select_related('student')
        unique_students = set() 

        student_data = []
        for mark in marks:
            if mark.student.id not in unique_students:
                unique_students.add(mark.student.id)
                student_data.append({
                    'mark_id': mark.id,
                    'name': mark.student.get_full_name(),
                })

        context['materia'] = materia
        context['student_data'] = student_data

        return context

#evoluvion de un studendiate dado
def evolution(request, materia_id, student_id):
    materia = get_object_or_404(Materia, id=materia_id)
    teacher = materia.teacher.get_full_name()
    student = get_object_or_404(User, pk=student_id)
    
    student_data = {
        'id': student.id,
        'username': student.username,
        'email': student.email,
        'first_name': student.first_name,
        'last_name': student.last_name,
        'profile_image_url': request.build_absolute_uri(student.profile.image.url),  # Agregar esta línea
    }
  
    marks = Mark.objects.filter(materia=materia, student=student_id)
    marks_data = [
        {
            'mark_1': item.mark_1,
            'mark_2': item.mark_2,
            'mark_3': item.mark_3,
            'average': item.average
        }
        for item in marks
    ]

    evolution_data = {
        'teacher': teacher,
        'materiaName': materia.name,
        'marks': marks_data,
        'student_data': student_data
    }
    
    # Ruta absoluta para la imagen de la marca de agua
    watermark_url = request.build_absolute_uri(staticfiles_storage.url('images/ADV.jpg'))
    evolution_data['watermark_url'] = watermark_url
    
    # Verificar si el código y la fecha/hora están en la sesión
    if 'unique_code' not in request.session or 'generated_at' not in request.session:
        # Generar código único
        unique_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
        request.session['unique_code'] = unique_code
        # Generar fecha y hora actuales
        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        request.session['generated_at'] = generated_at
    else:
        unique_code = request.session['unique_code']
        generated_at = request.session['generated_at']

    # Añadir código único y fecha/hora al contexto
    evolution_data['unique_code'] = unique_code
    evolution_data['generated_at'] = generated_at
    
    if request.GET.get('generate_pdf') == 'true':
        template = get_template('profile/evolution_pdf.html')
        html = template.render(evolution_data)

        response = HttpResponse(content_type='application/pdf')
        filename = f"{materia.name.replace(' ', '-')}_{student.username}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        html = HTML(string=html)
        result = html.write_pdf(encoding='utf-8')
        response.write(result)
        
        return response

    else:
        return JsonResponse(evolution_data, safe=False)

#vista para activar o desactivar el boton de caligficaciones
class ToggleAddNotesView(View):
    def post(self, request, *args, **kwargs):
        if request.user.groups.filter(name='administrativos').exists():
            is_active = request.POST.get('is_active') == 'true'
            config, created = GlobalConfig.objects.get_or_create(id=1)
            config.allow_add_notes = is_active
            config.save()
            messages.success(request, 'La configuración ha sido actualizada.')
        else:
            messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect(reverse('profile'))
@csrf_exempt
def update_add_notes_status(request):
    if request.method == "POST":
        data = json.loads(request.body)
        allow_add_notes = data.get("allow_add_notes", False)
        
        # Actualizar el estado en la base de datos
        request.user.profile.allow_add_notes = allow_add_notes
        request.user.profile.save()

        return JsonResponse({"success": True})
    return JsonResponse({"success": False})

#mostarar lista de alumnos y notas a los profesores sin accion de poner nota
@add_group_name_to_context
class StudentListMarkViewsacction(TemplateView, LoginRequiredMixin, UserPassesTestMixin):
    template_name = 'student_2_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        materia_id = self.kwargs['materia_id']
        materia = get_object_or_404(Materia, id=materia_id)
    
        context['teacher_name'] = materia.teacher.get_full_name()
        context['grading_enabled'] = materia.activate_grading
        
        # Obtener las notas y los estudiantes relacionados, ordenados por apellido
        marks = Mark.objects.filter(materia=materia).order_by('student__last_name')
        
        student_data = {}
        for mark in marks:
            student = get_object_or_404(User, id=mark.student_id)
            if mark.student_id not in student_data:
                student_data[mark.student_id] = {
                    'mark_id': mark.id,
                    #'name': student.get_full_name(),
                    'last_name': student.last_name,
                    'first_name': student.first_name,
                    'mark_1': mark.mark_1,
                    'mark_2': mark.mark_2,
                    'mark_3': mark.mark_3,
                    'average': mark.average,
                }

        context['materia'] = materia
        context['student_data'] = list(student_data.values()) 
        return context

#para buscar profesores
@add_group_name_to_context
class AdminProfesoresView(TemplateView):
    template_name = 'admin-profesores.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Mostrar todos los profesores al cargar la página
        context['user_profiles'] = User.objects.filter(groups__name='profesores')
        return context
def search_profesores(request):
    query = request.GET.get('q', '')
    user_profiles = []
    
    if query:
        # Filtrar los profesores por nombre o apellido
        user_profiles = User.objects.filter(
            Q(first_name__icontains=query) | Q(last_name__icontains=query),
            groups__name='profesores'
        ).values('first_name', 'last_name', 'username')
    else:
        # Si no hay búsqueda, devolver todos los profesores
        user_profiles = User.objects.filter(groups__name='profesores').values('first_name', 'last_name', 'username')

    return JsonResponse(list(user_profiles), safe=False)

#para ver cursos en admin list
@add_group_name_to_context
class CoursesView_list(TemplateView,LoginRequiredMixin,UserPassesTestMixin):
    template_name = 'admin-courses.html'  # Asegúrate de que el nombre del template es correcto

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        courses = Course.objects.all().order_by('-id')
        # Añade un print para verificar los cursos
        print("Cursos en el contexto:", courses)
        
        student = self.request.user if self.request.user.is_authenticated else None
        
        for item in courses:
            if student:
                registration = Registration.objects.filter(course=item, student=student).first()
                item.is_enrolled = registration is not None
            else:
                item.is_enrolled = False

            # La cantidad de estudiantes ya no se requiere según tu solicitud
            # enrollment_count = Registration.objects.filter(course=item).count()
            # item.enrollment_count = enrollment_count
        
        context['courses'] = courses
        return context
    #materias list
@add_group_name_to_context
class MateriaListView(TemplateView,LoginRequiredMixin,UserPassesTestMixin):
    template_name = 'admin-materias.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['materias'] = Materia.objects.all().order_by('name')  # Obtener todas las materias ordenadas por nombre
        return context

#para poder ver detalles de materia y curso
@add_group_name_to_context
class CourseDetailView(UserPassesTestMixin,LoginRequiredMixin, DetailView):
    model = Course
    template_name = 'course_detail.html'
    context_object_name = 'course'
    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name='secretarias').exists()
    def handle_no_permission(self):
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Obtener las materias relacionadas con el curso y ordenarlas por id (orden de creación)
        context['materias'] = self.object.materias.all().order_by('id')  # Orden ascendente por id
        return context


def materia_detail(request, materia_id):
    materia = get_object_or_404(Materia, id=materia_id)
    estudiantes = materia.registration_set.all()
    # Pasar el objeto materia con el profesor
    return render(request, 'materia_detail.html', {
        'materia': materia,
        'estudiantes': estudiantes,
        'teacher': materia.teacher 
    })

    #cambair contraseña de usuario
@add_group_name_to_context
class ProfilePasswordChangeView(PasswordChangeView):
    template_name = 'profile/change_password.html'
    success_url = reverse_lazy('profile')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['password_changed'] = self.request.session.get('password_changed', False)
        return context

    def form_valid(self, form):
        # Actualizar el campo created_by_admin del modelo Profile
        profile = Profile.objects.get(user=self.request.user)
        profile.create_by_admin = False
        profile.save()

        messages.success(self.request, 'Cambio de contraseña exitoso')
        update_session_auth_hash(self.request, form.user)
        self.request.session['password_changed'] = True
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(
            self.request,
            'Hubo un error al momento de intentar cambiar la contraseña: {}.'.format(
                form.errors.as_text()
            )
        )
        return super().form_invalid(form)

# AGREGAR UN NUEVO USUARIO
from django.db import IntegrityError

@add_group_name_to_context
class AddUserView(UserPassesTestMixin, LoginRequiredMixin, CreateView):
    model = User
    form_class = UserCreationForm
    template_name = 'add_user.html'
    success_url = '/profile/'

    def form_invalid(self, form):
        # Pasar los errores del formulario al contexto
        return self.render_to_response(self.get_context_data(form=form, errors=form.errors))

    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name='secretarias').exists()

    def handle_no_permission(self):
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        groups = Group.objects.all()
        courses = Course.objects.all()
        singular_groups = [plural_to_singular(group.name).capitalize() for group in groups]
        context['groups'] = zip(groups, singular_groups)
        context['courses'] = courses
        context['errors'] = kwargs.get('errors', None)  # Agregar los errores al contexto

        return context

    def form_valid(self, form):
        group_id = self.request.POST['group']
        course_id = self.request.POST.get('course')
        group = Group.objects.get(id=group_id)

        try:
            # Guardar el usuario
            user = form.save(commit=False)
            user.set_password('contraseña')  
            if group_id != '1':
                user.is_staff = True
            user.save()

            # Asignar grupo al usuario
            user.groups.clear()
            user.groups.add(group)

            # Inscribir al estudiante en el curso, si corresponde
            if group.name == 'estudiantes' and course_id:
                course = Course.objects.get(id=course_id)
                Registration.objects.create(course=course, student=user)

            messages.success(self.request, f'Usuario {user.username} creado exitosamente.')
            return super().form_valid(form)
        
        except IntegrityError:
            messages.error(self.request, 'El nombre de usuario ya está en uso. Por favor, elige otro.')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('add_user')


# Vista adicional para obtener materias relacionadas con un curso específico
class GetMateriasByCourseView(View):
    def get(self, request, *args, **kwargs):
        course_id = request.GET.get('course_id')
        materias = Materia.objects.filter(courses__id=course_id)
        data = [{"id": materia.id, "name": materia.name} for materia in materias]
        return JsonResponse(data, safe=False)


# LOGIN PERSONALIZADO
@add_group_name_to_context
class CustomLoginView(LoginView):
    def form_valid(self, form):
        response = super().form_valid(form)

        # Acceder al perfil del usuario
        profile = self.request.user.profile

        # Verificamos el valor del campo created_by_admin
        if profile.create_by_admin:
            messages.info(self.request, 'BIENVENIDO, Cambie su contraseña ahora !!!')
            response['Location'] = reverse_lazy('profile_password_change')
            response.status_code = 302
        return response
    def get_success_url(self):
        return super().get_success_url()

# VISUALIZACION DEL PERFIL DE UN USUARIO
@add_group_name_to_context
class UserDetailsView(LoginRequiredMixin,UserPassesTestMixin, DetailView):
    model = User
    template_name = 'user_details.html'
    context_object_name = 'user_profile'
    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name='secretarias').exists()
    def handle_no_permission(self):
        return redirect('error')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.get_object()
        group_id,group_name, group_name_singular, color = get_group_and_color(user)
   
    
        #obtener todos los grupos
        groups = Group.objects.all()
        singular_names = [plural_to_singular(group.name).capitalize() for group in groups]
        groups_ids = [group.id for group in groups]
        singular_groups = zip(singular_names, groups_ids)
        context['group_id_user'] = group_id
        context['group_name_user'] = group_name
        context['group_name_singular_user'] = group_name_singular
        context['color_user'] = color
        context['singular_groups']=singular_groups
         
         # Obtener el grupo del usuario
        if user.groups.filter(name='profesores').exists():
            context['group_name'] = 'profesores'
            assigned_materias = Materia.objects.filter(teacher=user).order_by('name')
            context['assigned_materias'] = assigned_materias

        elif user.groups.filter(name='estudiantes').exists():
            context['group_name'] = 'estudiantes'
            student_id=user.id
            registrations = Registration.objects.filter(student=user)
            enrolled_materias_set = set()
            for registration in registrations:
                materias = registration.course.materias.all()
                enrolled_materias_set.update(materias)
            enrolled_materias = list(enrolled_materias_set)
            context['student_id']=student_id
            context['enrolled_materias'] = enrolled_materias

        elif user.groups.filter(name='secretarias').exists():
            all_courses = Course.objects.prefetch_related('materias').all()
            context['all_courses'] = all_courses

        elif user.groups.filter(name='administrativos').exists():
            context['group_name'] = 'administrativos'
            all_users = User.objects.all()
            all_groups = Group.objects.all()
            user_profiles = []
            for user in all_users:
                profile = user.profile
                user_groups = user.groups.all()
                user_profiles.append({
                    'user': user,
                    'groups': user_groups,
                    'profile': profile
                })
                all_courses = Course.objects.all()
            context['all_courses'] = all_courses
            context['user_profiles'] = user_profiles
            context['all_groups'] = all_groups
        
        return context 

# GRABACION DE LOS DATOS DE UN USUARIO
def superuser_edit(request, user_id):
    if not request.user.is_superuser:
        return redirect('error')

    user = User.objects.get(pk=user_id)
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=user)
        profile_form = ProfileForm(request.POST, request.FILES, instance=user.profile)
        group = request.POST.get('group')

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            user.groups.clear()
            user.groups.add(group)
            return redirect('user_details', pk=user.id)
    else:
        user_form = UserForm(instance=user)
        profile_form = ProfileForm(instance=user.profile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form
    }
    return render(request, 'profile/user_details.html', context)


#genera pdf profesor
def generate_pdf_student_list(request, materia_id):
    materia = get_object_or_404(Materia, id=materia_id)
    marks = Mark.objects.filter(materia=materia)

    student_data = {}
    for mark in marks:
        student = get_object_or_404(User, id=mark.student_id)
        
        full_name = student.get_full_name()
        name_parts = full_name.split()
        
        if len(name_parts) >= 3:
            last_name = f"{name_parts[-2]} {name_parts[-1]}"
            first_name = " ".join(name_parts[:-2])
        elif len(name_parts) == 2:
            first_name = name_parts[0]
            last_name = name_parts[1]
        else:
            first_name = full_name
            last_name = ""

        if mark.student_id not in student_data:
            student_data[mark.student_id] = {
                'first_name': first_name,
                'last_name': last_name,
                'mark_1': mark.mark_1,
                'mark_2': mark.mark_2,
                'mark_3': mark.mark_3,
                'average': mark.average,
            }

    student_data_list = sorted(student_data.values(), key=lambda x: x['last_name'].lower())

    
    # Obtener la fecha y hora actuales
    date_time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # Construir la URL absoluta para la imagen de la marca de agua
    watermark_url = request.build_absolute_uri(static('images/ADV.jpg'))

    # Renderizar el template en HTML
    html_string = render_to_string('student_list_pdf.html', {
        'materia': materia,
        'student_data': student_data_list,
        'teacher': materia.teacher.get_full_name(),
        'watermark_url': watermark_url,  # Pasar la URL absoluta
        'date_time': date_time_now,  # Pasar la fecha y hora actuales

        
    })
    
    # Generar el PDF con WeasyPrint
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="listado_alumnos_{materia.name}.pdf"'

    HTML(string=html_string).write_pdf(response)
    
    return response


#genera excel admin
def generate_course_excel(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    materias = course.materias.all().order_by('id')

    # Obtener la lista de estudiantes y sus notas
    students = []
    for registration in Registration.objects.filter(course=course):
        student = registration.student
        marks = Mark.objects.filter(student=student, materia__in=materias)
        students.append({
            'student': student,
            'marks': marks
        })

    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Curso {course.name}"

    header_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    bold_font = Font(bold=True)

    ws["A1"] = f"Curso: {course.name}"
    ws["A1"].font = bold_font

    row_num = 2
    ws.append(["Estudiante"]) 

    col_start = 2  # Empezar en la columna B
    for materia in materias:
        ws.merge_cells(start_row=row_num, start_column=col_start, end_row=row_num, end_column=col_start + 3)
        ws.cell(row=row_num, column=col_start).value = materia.name
        ws.cell(row=row_num, column=col_start).font = bold_font

        # Escribir las columnas de notas en la fila siguiente
        ws.cell(row=row_num + 1, column=col_start).value = "1"
        ws.cell(row=row_num + 1, column=col_start + 1).value = "2"
        ws.cell(row=row_num + 1, column=col_start + 2).value = "3"
        ws.cell(row=row_num + 1, column=col_start + 3).value = "Promedio"

        # Aplicar el estilo de cabecera
        for col in range(col_start, col_start + 4):
            ws.cell(row=row_num, column=col).fill = header_fill
            ws.cell(row=row_num + 1, column=col).fill = header_fill
            ws.cell(row=row_num + 1, column=col).font = bold_font

        col_start += 4  # Moverse a la siguiente sección de columnas para la próxima materia

    # Añadir el encabezado para la columna de "Promedio Final"
    ws.cell(row=row_num, column=col_start).value = "Promedio Final"
    ws.cell(row=row_num, column=col_start).font = bold_font
    ws.cell(row=row_num, column=col_start).fill = header_fill

    # Escribir los datos de los estudiantes
    row_num += 2  # Empezar en la fila después del encabezado de materias
    for student in students:
        row = [student['student'].get_full_name()]  # Nombre completo del estudiante
        averages = []  # Lista para almacenar los promedios de cada materia para el estudiante

        for materia in materias:
            mark = student['marks'].filter(materia=materia).first()
            if mark:
                row.extend([mark.mark_1, mark.mark_2, mark.mark_3, mark.average])
                if mark.average is not None:  # Asegúrate de que el promedio no sea None
                    averages.append(mark.average)  # Añadir el promedio de la materia a la lista de promedios
            else:
                row.extend(["", "", "", ""])

        # Calcular el promedio de los promedios de las materias
        if averages:
            final_average = sum(averages) / len(averages)  # Promedio de los promedios
            row.append(final_average)
        else:
            row.append("")

        ws.append(row)

    # Crear una respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{course.name}.xlsx"'

    # Guardar el libro de trabajo en la respuesta
    wb.save(response)

    return response


#boton de acerca del colegio
@add_group_name_to_context
class UnidadEducativaView(TemplateView):
    template_name = 'unidad_educativa.html'

#register atenndence teacher
@add_group_name_to_context
class RegisterEntryOrExitView(TemplateView):
    template_name = 'register_entry_or_exit.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = AttendanceForm()  # Inicializa el formulario
        return context

    def post(self, request, *args, **kwargs):
        form = AttendanceForm(request.POST)
        if form.is_valid():
            ci = form.cleaned_data['ci']
            teacher = get_object_or_404(CITeacher, ci=ci)

            if 'entry' in request.POST:
                # Registro de entrada
                attendance, created = TeacherAttendance.objects.get_or_create(teacher=teacher, entry_time__isnull=True)
                if not attendance.entry_time:  # Solo si no ha registrado la entrada
                    attendance.entry_time = timezone.now()
                    attendance.save()
                    message = 'Entrada registrada con éxito'
                else:
                    message = 'Ya has registrado la entrada anteriormente.'

            elif 'exit' in request.POST:
                # Registro de salida
                attendance = TeacherAttendance.objects.filter(teacher=teacher, exit_time__isnull=True).last()
                if attendance and attendance.entry_time and not attendance.exit_time:  # Solo si no ha registrado la salida
                    attendance.exit_time = timezone.now()
                    attendance.save()
                    message = 'Salida registrada con éxito'
                else:
                    message = 'Primero debes registrar la entrada antes de registrar la salida.'

            return render(request, 'attendance_success.html', {'message': message})

        # En caso de que el formulario no sea válido, vuelve a mostrar el formulario con errores
        return render(request, self.template_name, {'form': form})

#tempalte para ver lista de entrada y salida
@add_group_name_to_context
class attendance_prof(UserPassesTestMixin, ListView):
    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name='secretarias').exists()

    def handle_no_permission(self):
        return redirect('error')

    model = TeacherAttendance
    template_name = 'attendance_prof.html'
    context_object_name = 'attendances'
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = super().get_queryset()
    
    # Obtener las fechas del formulario
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        print(f"Start Date: {start_date}, End Date: {end_date}")  

    # Verificar si las fechas están presentes
        if start_date and end_date:
            try:
                start_date = timezone.datetime.fromisoformat(start_date)
                end_date = timezone.datetime.fromisoformat(end_date)
                queryset = queryset.filter(created_at__range=[start_date, end_date + timezone.timedelta(days=1)])
                
                print(f"Filtered Queryset Count: {queryset.count()}")  # Debugging print
            except ValueError:
                # Manejo de error si las fechas no se pueden convertir
                queryset = queryset.none()
        
        return queryset


#csv
@add_group_name_to_context
class AddUsersFromCSVView(UserPassesTestMixin, LoginRequiredMixin, TemplateView):
    template_name = 'add_users_from_csv.html'

    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name='secretarias').exists()

    def handle_no_permission(self):
        return redirect('error')

    def post(self, request):
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            messages.error(request, "Por favor, selecciona un archivo CSV.")
            return redirect('add_users_from_csv')

        try:
            try:
                data = csv_file.read().decode('utf-8')
            except UnicodeDecodeError:
                messages.error(request, "Error de codificación. Asegúrate de que el archivo esté en formato UTF-8.")
                return redirect('add_users_from_csv')

            csv_reader = csv.DictReader(data.splitlines(), delimiter=';')
            headers = csv_reader.fieldnames
            if headers:
                print(f"Cabeceras en el archivo CSV: {headers}")
            else:
                messages.error(request, "El archivo CSV no tiene cabeceras.")
                return redirect('add_users_from_csv')

            for index, row in enumerate(csv_reader, start=1):
                print(f"Procesando fila {index}: {row}")

                username = row.get('username')
                email = row.get('email')
                group_id = row.get('group')
                first_name = row.get('firstname')
                last_name = row.get('lastname')
                course_name = row.get('course')

                if not all([username, email, group_id, first_name, last_name, course_name]):
                    messages.error(request, f"Faltan datos en la fila {index}. {row}")
                    continue

                try:
                    course = Course.objects.get(name__iexact=course_name.strip())
                except Course.DoesNotExist:
                    messages.error(request, f"El curso '{course_name}' no existe en la fila {index}.")
                    continue

                # Comprobar si el usuario ya existe
                user, created = User.objects.get_or_create(
                    username=username,
                    defaults={
                        'email': email,
                        'first_name': first_name,
                        'last_name': last_name,
                        'password': make_password('contraseña')
                    }
                )

                if not created:
                    messages.warning(request, f"El usuario '{username}' ya existe. Se omite la fila {index}.")
                    continue

                try:
                    group = Group.objects.get(id=group_id)
                    user.groups.add(group)
                except Group.DoesNotExist:
                    messages.error(request, f"El grupo con ID {group_id} no existe en la fila {index}.")
                    user.delete()
                    continue

                try:
                    if group.name.lower() == 'estudiantes':
                        # Verificar si el estudiante ya está inscrito en el curso
                        registration, registration_created = Registration.objects.get_or_create(student=user, course=course)

                        if registration_created:
                            # Solo crear las notas si la inscripción es nueva
                            for materia in course.materias.all():
                                Mark.objects.get_or_create(student=user, materia=materia)
                        else:
                            messages.warning(request, f"El estudiante '{user.username}' ya está inscrito en el curso '{course_name}'. Se omite la fila {index}.")
                except Exception as e:
                    messages.error(request, f"Error al inscribir al estudiante en el curso '{course_name}' en la fila {index}: {str(e)}")
                    user.delete()
                    continue

            messages.success(request, "Usuarios agregados correctamente.")
            return redirect('add_users_from_csv')

        except Exception as e:
            messages.error(request, f"Hubo un error al procesar el archivo CSV: {str(e)}")
            return redirect('add_users_from_csv')

        
#elimninar usuarios
@add_group_name_to_context
class DeleteUserView(UserPassesTestMixin, LoginRequiredMixin, DeleteView):
    model = User
    template_name = 'delete_user.html'
    success_url = '/profile/'  # Redirigir a la página de perfil o la lista de usuarios después de la eliminación

    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name='secretarias').exists()

    def handle_no_permission(self):
        return redirect('error')

#horario
@add_group_name_to_context
class HorarioView(UserPassesTestMixin, LoginRequiredMixin, TemplateView):
    template_name = 'horario/horario.html'

    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name='secretarias').exists()

    def handle_no_permission(self):
        return redirect('error')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['courses'] = Course.objects.all()
        context['days_of_week'] = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
        context['time_slots'] = ["1", "2", "3", "4", "5", "6", "7", "8"]

        # Obtener materias basadas en el curso seleccionado
        selected_course_id = self.request.GET.get('course')
        if selected_course_id:
            selected_course = Course.objects.filter(id=selected_course_id).first()
            if selected_course:
                 context['classes'] = selected_course.materias.all()

            else:
                context['classes'] = []
        else:
            context['classes'] = []

        context['draggable_enabled'] = True
        return context

#para subir pdf de horario
def upload_schedule(request):
    if request.method == 'POST':
        # Asegurarse de que se reciban múltiples archivos
        files = request.FILES.getlist('file')  # 'file' es el nombre del campo de entrada
        if files:
            for file in files:
                schedule = Schedule(file=file)
                schedule.name = file.name  # Guardar el nombre del archivo PDF
                schedule.save()
            messages.success(request, f"{len(files)} archivos subidos exitosamente.")
            return redirect('schedule_list')
        else:
            messages.error(request, "No se seleccionaron archivos.")
            return redirect('schedule_list')

    return render(request, 'upload_schedule.html')
#ver horarios admin
@add_group_name_to_context
class ScheduleListView(UserPassesTestMixin, ListView):
    model = Schedule
    template_name = 'schedule_list.html'
    context_object_name = 'schedules'

    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name='secretarias').exists()
    def handle_no_permission(self):
        return redirect('error')
    def get_queryset(self):
        return Schedule.objects.all()


    def post(self, request, *args, **kwargs):
        # Verificar si se quiere eliminar un archivo
        file_id = request.POST.get('delete_file_id')
        if file_id:
            schedule = get_object_or_404(Schedule, id=file_id)
            schedule.delete()
            messages.success(request, f"El archivo {schedule.name} ha sido eliminado exitosamente.")
            return redirect('schedule_list')  # Redirigir después de eliminar el archivo

        return HttpResponseRedirect(self.get_success_url())  # Si no se eliminó nada, redirigir

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

#automatiza el horario view
def generate_pdf_from_html(html_content):
    # Convierte el HTML a un archivo PDF
    pdf_file = BytesIO()
    HTML(string=html_content).write_pdf(pdf_file)
    return pdf_file.getvalue()  # Devuelve el contenido binario del PDF

@csrf_exempt
def generate_and_save_schedule_pdf(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            course_id = data.get("course_id")
            pdf_base64 = data.get("pdf_data")  # Recibe el PDF en formato base64

            if not course_id or not pdf_base64:
                return JsonResponse({"error": "Datos incompletos"}, status=400)

            # Decodifica el contenido base64 del PDF
            pdf_content = base64.b64decode(pdf_base64)

            # Obtener el nombre del curso
            course = get_object_or_404(Course, id=course_id)
            course_name = course.name.replace(" ", "_")

            # Guardar el archivo PDF en el modelo Schedule
            file_name = f"{course_name}.pdf"
            schedule = Schedule(
                name=file_name,
                file=ContentFile(pdf_content, name=file_name),
            )
            schedule.save()

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    else:
        return JsonResponse({"error": "Método no permitido"}, status=405)

#acivar notas estudnet
class ToggleViewEvolutionView(View):
    def post(self, request, *args, **kwargs):
        if request.user.groups.filter(name='administrativos').exists(): 
            is_active = request.POST.get('is_active') == 'true'
            config, created = GlobalConfig.objects.get_or_create(id=1)
            config.allow_view_evolution = is_active  
            config.save()
            messages.success(request, 'La configuración para Consultar Evolución ha sido actualizada.')
        else:
            messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect(reverse('profile'))

#ver horario estudent
@add_group_name_to_context 
class ScheduleDetailView(View):
    def get(self, request, *args, **kwargs):
        student = request.user
        registration = Registration.objects.filter(student=student, enabled=True).first()

        if registration:
            course = registration.course
            schedules = Schedule.objects.filter(name__icontains=course.name)

            if schedules.exists():
                context = {
                    'course': course,
                    'schedules': schedules
                }
                context.update(self.extra_context) 

                return render(request, 'horario_student/schedule_detail.html', context)
            else:
                context = {
                    'course': course
                }
                context.update(self.extra_context) 
                return render(request, 'horario_student/no_schedule_found.html', context)
        else:
            context = {
                'student': student
            }
            context.update(self.extra_context) 
            return render(request, 'horario_student/no_registration.html', context)
#editar materias 
@add_group_name_to_context
class MateriaEditView(View):
    template_name = 'edit_materia.html'

    def get(self, request):
        courses = Course.objects.all()
        selected_course = request.GET.get('course')
        selected_materia = request.GET.get('materia')

        materias = Materia.objects.filter(courses__id=selected_course) if selected_course else None
        materia = get_object_or_404(Materia, id=selected_materia) if selected_materia else None

        form = EditMateriaForm(instance=materia) if materia else None

        # Combinamos el contexto proporcionado por el decorador con el contexto original
        context = {
            'courses': courses,
            'materias': materias,
            'form': form,
        }
        if hasattr(self, 'extra_context'):
            context.update(self.extra_context)  # Añadimos el contexto extra proporcionado por el decorador

        return render(request, self.template_name, context)

    def post(self, request):
        selected_course = request.POST.get('course')
        selected_materia = request.POST.get('materia')
        materia = get_object_or_404(Materia, id=selected_materia)

        form = EditMateriaForm(request.POST, instance=materia)
        if form.is_valid():
            form.save()
            messages.success(request, 'La materia se ha actualizado correctamente.')
            return redirect('materia_edit')  # Cambia esto a la URL correcta

        materias = Materia.objects.filter(courses__id=selected_course)

        # Combinamos el contexto proporcionado por el decorador con el contexto original
        context = {
            'courses': Course.objects.all(),
            'materias': materias,
            'form': form,
        }
        if hasattr(self, 'extra_context'):
            context.update(self.extra_context)  # Añadimos el contexto extra proporcionado por el decorador

        return render(request, self.template_name, context)
    
#template de fin de año


#generar csv de cada curso con estudiantes aprobados en almenos materia
def generate_course_preview(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    materias = course.materias.all().order_by('id')
    
    eligible_students = []
    ineligible_students = []
    
    for registration in Registration.objects.filter(course=course):
        student = registration.student
        marks = Mark.objects.filter(student=student, materia__in=materias)
        
        if any(mark.average > 50 for mark in marks if mark.average is not None):
            eligible_students.append({
                'username': student.username,
                'email': student.email,
                #'group': 1,
                'firstname': student.first_name,
                'lastname': student.last_name,
                'course': course.name
            })
        else:
            ineligible_students.append({
                'username': student.username,
                'email': student.email,
                #'group': 1,
                'firstname': student.first_name,
                'lastname': student.last_name,
                'course': course.name
            })
    
    return JsonResponse({
        'eligible_students': eligible_students,
        'ineligible_students': ineligible_students
    })

def download_course_csv(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    materias = course.materias.all().order_by('id')
    
    # Filter students with an average mark above 51 in each subject
    eligible_students = []
    for registration in Registration.objects.filter(course=course):
        student = registration.student
        user_profile = student.profile
        
        # Check if the student has an average mark > 51 in each subject
        marks = Mark.objects.filter(student=student, materia__in=materias)
        if all(mark.average > 50 for mark in marks if mark.average is not None):
            eligible_students.append({
                'username': student.username,
                'email': student.email,
                'group': 1,
                'firstname': student.first_name,
                'lastname': student.last_name,
                'course': course.name
            })

    # Create a CSV response with a semicolon delimiter
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{course.name}_aprobados.csv"'
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['username', 'email', 'group', 'firstname', 'lastname', 'course'])
    
    # Write eligible students to CSV
    for student in eligible_students:
        writer.writerow([student['username'], student['email'], student['group'], 
                         student['firstname'], student['lastname'], student['course']])
    
    return response

#student history
def student_detail(request, student_id):
    student = get_object_or_404(User, id=student_id)
    history = StudentHistory.objects.filter(student=student).order_by('-completion_date')
    
    return render(request, 'student_detail.html', {
        'student': student,
        'history': history,
    })
#siguente curso
def move_student_with_history(student, current_course, new_course):
    """
    Mueve un estudiante de un curso a otro, archivando su historial y gestionando sus relaciones con notas y materias.
    """
    try:
        with transaction.atomic():
            # 1. Obtener notas actuales del estudiante en el curso anterior
            marks = Mark.objects.filter(student=student, materia__in=current_course.materias.all())

            # 2. Calcular promedio final del curso anterior
            final_average = marks.aggregate(average=Avg('average'))['average'] or 0

            # 3. Crear el historial del curso anterior
            history = StudentHistory.objects.create(
                student=student,
                course=current_course,
                enrollment_date=Registration.objects.get(student=student, course=current_course).created_at,  # Si tienes este campo
                completion_date=date.today(),
                is_approved=final_average >= 51,  # Aprobado si cumple el promedio
                final_average=final_average,
                detailed_grades={mark.materia.name: {
                    'mark_1': mark.mark_1,
                    'mark_2': mark.mark_2,
                    'mark_3': mark.mark_3,
                    'average': mark.average
                } for mark in marks},
                comment=f"Promedio final: {final_average}. Transición a {new_course.name}."
            )

            # 4. Eliminar las notas del curso anterior
            marks.delete()

            # 5. Actualizar la inscripción del estudiante al nuevo curso
            registration = Registration.objects.get(student=student, course=current_course)
            registration.course = new_course
            registration.save()

            # 6. Crear nuevas notas para las materias del nuevo curso
            new_materias = new_course.materias.all()
            for materia in new_materias:
                Mark.objects.get_or_create(student=student, materia=materia)

            return True
    except Exception as e:
        print(f"Error al mover al estudiante con historial: {e}")
        return False


#modal de estudiantes del curso
def get_course_students(request, course_id):
    """
    Devuelve los estudiantes registrados en un curso específico como JSON.
    """
    course = get_object_or_404(Course, id=course_id)
    
    # Obtener inscripciones activas de estudiantes para el curso
    registrations = Registration.objects.filter(course=course, enabled=True).select_related('student')

    # Crear una lista de estudiantes con datos básicos
    student_data = [
        {
            'name': f"{reg.student.first_name} {reg.student.last_name}",
            #'email': reg.student.email,
        }
        for reg in registrations
    ]

    return JsonResponse({'students': student_data})

#movel student course
@csrf_exempt
def move_students_to_course(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_ids = data.get('student_ids', [])
            target_course_id = data.get('target_course_id')
            current_course_id = data.get('current_course_id')

            # Validar cursos
            current_course = get_object_or_404(Course, id=current_course_id)
            target_course = get_object_or_404(Course, id=target_course_id)

            # Validar estudiantes
            registrations = Registration.objects.filter(
                course=current_course, student_id__in=student_ids
            )
            if not registrations.exists():
                return JsonResponse({'error': 'No se encontraron estudiantes válidos para mover.'}, status=400)

            # Actualizar inscripciones
            registrations.update(course=target_course)
            return JsonResponse({'message': 'Estudiantes movidos exitosamente.'}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido.'}, status=405)

def course_detail(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    all_courses = Course.objects.exclude(id=course_id)  # Excluye el curso actual
    materias = course.materias.all()

    context = {
        'course': course,
        'materias': materias,
        'all_courses': all_courses
    }
    return render(request, 'course_detail.html', context)

def get_available_courses(request, current_course_id):
    courses = Course.objects.exclude(id=current_course_id)  # Excluye el curso actual
    course_data = [{'id': course.id, 'name': course.name} for course in courses]
    return JsonResponse({'courses': course_data})

@csrf_exempt
def move_students(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            target_course_id = data.get('targetCourseId')
            student_ids = data.get('studentIds')

            print("Curso destino:", target_course_id)
            print("Estudiantes seleccionados:", student_ids)

            # Valida los datos
            if not target_course_id or not student_ids:
                return JsonResponse({'success': False, 'error': 'Datos incompletos.'}, status=400)

            # Realiza el movimiento de estudiantes
            # (Agrega aquí tu lógica para actualizar los estudiantes)

            return JsonResponse({'success': True})
        except Exception as e:
            print("Error interno:", e)
            return JsonResponse({'success': False, 'error': 'Error interno en el servidor.'}, status=500)
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)

#view move 3
@add_group_name_to_context
class MoveStudentView(View):
    template_name = 'profile/move_student.html'

    def get_context_data(self, **kwargs):
        course_id = self.kwargs['course_id']
        course = get_object_or_404(Course, id=course_id)

        students = User.objects.filter(groups__name='estudiantes', registration__course=course)

        courses = Course.objects.exclude(id=course_id)

        context = {
            'course': course,
            'students': students,
            'courses': courses,
        }
        if hasattr(self, 'extra_context'):
            context.update(self.extra_context)
        return context

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        student_ids = request.POST.getlist('students')  # Usar getlist para obtener múltiples IDs
        course_id = request.POST.get('course')

        if not student_ids or not course_id:
            messages.error(request, "Debes seleccionar al menos un estudiante y un curso.")
            return redirect(request.path)

        # Verificar que el curso existe
        new_course = get_object_or_404(Course, id=course_id)

        for student_id in student_ids:
            # Verificar que el estudiante existe
            student = get_object_or_404(User, id=student_id, groups__name='estudiantes')

            # Obtener la inscripción del estudiante
            registration = Registration.objects.filter(student=student).first()
            if registration:
                old_course = registration.course
                registration.course = new_course
                registration.save()

                # Manejo de las notas
                old_materias = old_course.materias.all()
                new_materias = new_course.materias.all()

                # Obtener los IDs de las materias
                old_materia_ids = old_materias.values_list('id', flat=True)
                new_materia_ids = new_materias.values_list('id', flat=True)

                # Determinar las materias que ya no pertenecen al nuevo curso
                materias_to_remove = set(old_materia_ids) - set(new_materia_ids)

                # Eliminar las notas asociadas a las materias que ya no están en el curso nuevo
                Mark.objects.filter(student=student, materia_id__in=materias_to_remove).delete()

                # Crear notas para las nuevas materias
                for materia in new_materias:
                    Mark.objects.get_or_create(student=student, materia=materia)

                messages.success(request, f"{student.get_full_name()} ha sido movido al curso {new_course.name}.")
            else:
                messages.error(request, f"No se encontró la inscripción del estudiante {student.get_full_name()}.")

        return redirect(request.path)

#libros
@add_group_name_to_context
class ManageBooksView(ListView):
    model = Book
    template_name = 'book/manage_books.html'
    context_object_name = 'books'

# Vista para añadir libros
@add_group_name_to_context
class AddBookView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'book/add_book.html'
    success_url = reverse_lazy('manage_books')

# Vista para editar libros
@add_group_name_to_context
class EditBookView(UpdateView):
    model = Book
    form_class = BookForm
    template_name = 'book/edit_book.html'
    success_url = reverse_lazy('manage_books')

# Vista para eliminar libros
@add_group_name_to_context
class DeleteBookView(DeleteView):
    model = Book
    template_name = 'book/delete_book.html'
    success_url = reverse_lazy('manage_books')

# Vista para que los estudiantes puedan ver los libros
@add_group_name_to_context
class ViewBooksView(ListView):
    model = Book
    template_name = 'book/view_books.html'
    context_object_name = 'books'
    
#paginacion estudent
